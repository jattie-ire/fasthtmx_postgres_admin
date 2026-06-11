"""FastAPI Admin Dashboard - Main Application"""
from fastapi import FastAPI, Request, Form
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from starlette.middleware.base import BaseHTTPMiddleware
from psycopg2.extras import RealDictCursor
from pathlib import Path
import subprocess
import json
import uuid
import threading
import time

# Import utility modules
from config import (
    config,
    SCRIPTS_CONFIG,
    TABLE_PERMISSIONS,
    get_table_permissions,
    can_user_view_table,
    can_user_edit_table,
    can_user_add_to_table,
    can_user_delete_from_table,
    can_user_view,
    can_user_edit,
    can_user_add,
    can_user_delete,
    is_user_admin,
    can_user_run_scripts,
    get_audit_trail_table_name,
    is_script_background,
)
from auth import kerberos_auth
from database import (
    get_db_connection,
    return_db_connection,
    get_available_tables,
    get_tables_with_display_names,
    get_table_columns,
    get_column_types,
    get_table_data,
    get_lookup_options,
    add_lookup_value,
    find_source_table,
    generate_lookups_for_table,
    get_user_permissions,
    ensure_user_exists,
    export_table_to_csv,
    export_table_to_excel,
    insert_imported_rows,
    get_disk_stats,
    initialize_connection_pool,
    create_audit_table,
    delete_user,
    log_audit_event,
    get_primary_key,
)
from templates import render_template

app = FastAPI()

# ============================================================================
# BACKGROUND SCRIPT JOB TRACKER
# ============================================================================

# In-memory job tracker for background scripts
_script_jobs = {}
_jobs_lock = threading.Lock()


def generate_execution_id():
    """Generate a unique execution ID for script jobs"""
    return str(uuid.uuid4())


def add_job(execution_id: str, script_name: str):
    """Add a new background job to the tracker"""
    with _jobs_lock:
        _script_jobs[execution_id] = {
            "status": "running",
            "script_name": script_name,
            "start_time": time.time(),
            "result": None,
        }


def update_job_result(execution_id: str, result: dict):
    """Update a job with completion result"""
    with _jobs_lock:
        if execution_id in _script_jobs:
            _script_jobs[execution_id]["status"] = "completed"
            _script_jobs[execution_id]["end_time"] = time.time()
            _script_jobs[execution_id]["result"] = result


def get_job_status(execution_id: str) -> dict:
    """Get the status of a job"""
    with _jobs_lock:
        if execution_id not in _script_jobs:
            return {"error": "Job not found", "status": "not_found"}
        
        job = _script_jobs[execution_id]
        return {
            "execution_id": execution_id,
            "status": job["status"],
            "script_name": job["script_name"],
            "start_time": job["start_time"],
            "end_time": job.get("end_time"),
            "result": job["result"],
        }


def get_active_scripts() -> list:
    """Get list of currently running scripts"""
    with _jobs_lock:
        return [
            {
                "execution_id": eid,
                "script_name": job["script_name"],
                "start_time": job["start_time"],
            }
            for eid, job in _script_jobs.items()
            if job["status"] == "running"
        ]


def cleanup_old_jobs():
    """Remove jobs completed more than 1 hour ago"""
    current_time = time.time()
    cutoff_time = current_time - 3600  # 1 hour
    
    with _jobs_lock:
        jobs_to_delete = [
            eid for eid, job in _script_jobs.items()
            if job["status"] == "completed" and job.get("end_time", 0) < cutoff_time
        ]
        for eid in jobs_to_delete:
            del _script_jobs[eid]

# ============================================================================
# HTTP CACHING MIDDLEWARE
# ============================================================================

class CacheHeaderMiddleware(BaseHTTPMiddleware):
    """Add cache headers for static assets"""
    async def dispatch(self, request: Request, call_next):
        response = await call_next(request)
        
        # Cache static files for 24 hours (86400 seconds)
        if request.url.path.startswith("/static/"):
            response.headers["Cache-Control"] = "public, max-age=86400"
        
        return response

app.add_middleware(CacheHeaderMiddleware)

# Mount static files
static_dir = Path(__file__).parent / "static"
static_dir.mkdir(exist_ok=True)
app.mount("/static", StaticFiles(directory=str(static_dir)), name="static")


# ============================================================================
# APP STARTUP
# ============================================================================

@app.on_event("startup")
async def startup_event():
    """Initialize database connection pool and create audit table on app startup"""
    print("Initializing application...")
    initialize_connection_pool()
    create_audit_table()
    print("✓ Application startup complete")


# ============================================================================
# STARTUP VALIDATION
# ============================================================================

def validate_table_permissions():
    """Validate that all columns in table_permissions exist"""
    if not TABLE_PERMISSIONS:
        return
    
    print("Validating table permissions...")
    for table_name, perms in TABLE_PERMISSIONS.items():
        editable_cols = perms.get("editable_columns", [])
        if not editable_cols:
            continue
        
        try:
            actual_columns = get_table_columns(table_name)
            if not actual_columns:
                raise ValueError(f"Table '{table_name}' not found in database")
            
            for col in editable_cols:
                if col not in actual_columns:
                    raise ValueError(f"Column '{col}' not found in table '{table_name}'. Available columns: {actual_columns}")
            
            print(f"✓ Table '{table_name}': permissions valid")
        except Exception as e:
            print(f"✗ ERROR validating table '{table_name}': {e}")
            raise


# Validate on startup
try:
    validate_table_permissions()
except Exception as e:
    print(f"FATAL: Table permissions validation failed: {e}")
    exit(1)


# ============================================================================
# AUTHENTICATION ROUTES
# ============================================================================


@app.get("/", response_class=HTMLResponse)
async def root(request: Request):
    """Home page"""
    username = request.cookies.get("username")
    if username:
        return await dashboard(request)
    
    from config import get_kerberos_login_text
    login_text = get_kerberos_login_text()
    html = render_template("login.html", {**login_text})
    return html


@app.get("/login", response_class=HTMLResponse)
async def get_login(request: Request):
    """Login page endpoint"""
    from config import get_kerberos_login_text
    login_text = get_kerberos_login_text()
    html = render_template("login.html", {**login_text})
    return html


@app.post("/login", response_class=HTMLResponse)
async def login(request: Request, username: str = Form(...), password: str = Form(...)):
    """Login endpoint"""
    from database import ensure_user_exists
    from config import get_kerberos_login_text
    
    if kerberos_auth(username, password):
        # Ensure user exists in fastapi_users table
        ensure_user_exists(username)
        
        response = RedirectResponse(url="/dashboard", status_code=303)
        response.set_cookie("username", username)
        return response
    
    login_text = get_kerberos_login_text()
    login_text["error"] = "Invalid credentials"
    html = render_template("login.html", {**login_text})
    return html


@app.get("/dashboard", response_class=HTMLResponse)
async def dashboard(request: Request):
    """Dashboard page - displays configured view"""
    username = request.cookies.get("username")
    if not username:
        return RedirectResponse(url="/", status_code=303)
    
    # Check user has view permission
    user_perms = get_user_permissions(username)
    if not can_user_view(user_perms):
        return render_template("error.html", {"error": "You do not have permission to view data", "username": username})
    
    # Get the view name from config
    view_name = config.get("dashboard", {}).get("display_view", "")
    data = []
    columns = []
    
    if view_name:
        # Fetch data from the view
        conn = get_db_connection()
        if conn:
            try:
                cursor = conn.cursor(cursor_factory=RealDictCursor)
                
                # First check if view exists
                cursor.execute("""
                    SELECT EXISTS (
                        SELECT 1 FROM information_schema.tables 
                        WHERE table_schema = 'public' 
                        AND table_name = %s
                    )
                """, (view_name,))
                exists = cursor.fetchone()['exists']
                
                if not exists:
                    # Check in all schemas
                    cursor.execute("""
                        SELECT table_schema, table_name FROM information_schema.tables 
                        WHERE table_name = %s
                        ORDER BY table_schema
                    """, (view_name,))
                    found = cursor.fetchall()
                    if found:
                        view_name = f"View found in schema: {found[0]['table_schema']}.{found[0]['table_name']} - Need to grant permissions or use full name"
                    else:
                        view_name = f"View '{view_name}' not found in database"
                else:
                    # Fetch the data
                    query = f"SELECT * FROM {view_name}"
                    print(f"DEBUG: Fetching dashboard view: {query}")
                    cursor.execute(query)
                    data = cursor.fetchall()
                    
                    # Get column names from cursor description or data
                    if cursor.description:
                        columns = [desc[0] for desc in cursor.description]
                    elif data:
                        columns = list(data[0].keys())
                    
                    print(f"Dashboard view '{view_name}': {len(data)} rows, {len(columns)} columns")
                    print(f"DEBUG: Columns: {columns}")
                
                cursor.close()
            except Exception as e:
                print(f"ERROR fetching dashboard view '{view_name}': {e}")
                import traceback
                traceback.print_exc()
                view_name = f"ERROR: {str(e)}"
            finally:
                conn.close()
    
    # Get display text from config
    display_text = config.get("dashboard", {}).get("display_text", "")
    display_description = config.get("dashboard", {}).get("display_description", "")
    
    user_perms = get_user_permissions(username)
    
    # Get age coloring configuration
    age_coloring_config = config.get("age_coloring", {})
    age_coloring_enabled = age_coloring_config.get("enabled", False)
    age_coloring_column = None
    
    if age_coloring_enabled:
        # Check if dashboard view is in the list of tables to apply age coloring
        age_coloring_tables = age_coloring_config.get("tables", [])
        should_apply = not age_coloring_tables or view_name in age_coloring_tables
        
        if should_apply:
            # Find first matching column in this table
            possible_columns = age_coloring_config.get("columns", ["last_updated", "created_date", "updated_at"])
            for col in possible_columns:
                if col in columns:
                    age_coloring_column = col
                    break
    
    html = render_template("dashboard.html", {
        "username": username,
        "page": "dashboard",
        "is_admin": is_user_admin(user_perms),
        "can_run_scripts": can_user_run_scripts(user_perms),
        "available_tables": [],
        "view_name": view_name,
        "display_text": display_text,
        "display_description": display_description,
        "columns": columns,
        "data": data,
        # Age coloring
        "age_coloring_enabled": age_coloring_enabled,
        "age_coloring_column": age_coloring_column,
        "max_age_days": age_coloring_config.get("max_age_days", 365),
    })
    return html


# ADMIN ENDPOINTS
# ============================================================================

@app.get("/admin", response_class=HTMLResponse)
async def admin_panel(request: Request):
    """Admin panel - manage users and config"""
    username = request.cookies.get("username")
    if not username:
        return RedirectResponse(url="/", status_code=303)
    
    user_perms = get_user_permissions(username)
    if not is_user_admin(user_perms):
        return render_template("error.html", {"error": "You do not have admin privileges", "username": username})
    
    html = render_template("admin.html", {
        "username": username,
        "page": "admin",
        "is_admin": True,
        "can_run_scripts": can_user_run_scripts(user_perms)
    })
    return html


@app.get("/api/admin/users")
async def api_get_users(request: Request):
    """Get list of all users from fastapi_users table"""
    username = request.cookies.get("username")
    if not username:
        return {"error": "Not authenticated"}
    
    user_perms = get_user_permissions(username)
    if not is_user_admin(user_perms):
        return {"error": "Admin privileges required"}
    
    from config import FASTAPI_USERS_TABLE
    
    conn = get_db_connection()
    if not conn:
        return {"error": "Database connection failed"}
    
    try:
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        query = f'SELECT "user", email, view, edit, add, delete, admin, run_scripts, export_data, import_data FROM {FASTAPI_USERS_TABLE} ORDER BY "user"'
        cursor.execute(query)
        users = cursor.fetchall()
        cursor.close()
        conn.close()
        return {"users": [dict(u) for u in users]}
    except Exception as e:
        print(f"ERROR fetching users: {e}")
        return {"error": str(e)}


@app.post("/api/admin/users/{user_to_edit}")
async def api_update_user(user_to_edit: str, request: Request):
    """Update user permissions"""
    username = request.cookies.get("username")
    if not username:
        return {"error": "Not authenticated"}
    
    user_perms = get_user_permissions(username)
    if not is_user_admin(user_perms):
        return {"error": "Admin privileges required"}
    
    from config import FASTAPI_USERS_TABLE
    
    form_data = await request.form()
    
    conn = get_db_connection()
    if not conn:
        return {"error": "Database connection failed"}
    
    try:
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # Parse boolean values from form
        view = form_data.get('view') == 'on' or form_data.get('view') == 'true'
        edit = form_data.get('edit') == 'on' or form_data.get('edit') == 'true'
        add = form_data.get('add') == 'on' or form_data.get('add') == 'true'
        delete = form_data.get('delete') == 'on' or form_data.get('delete') == 'true'
        admin = form_data.get('admin') == 'on' or form_data.get('admin') == 'true'
        run_scripts = form_data.get('run_scripts') == 'on' or form_data.get('run_scripts') == 'true'
        export_data = form_data.get('export_data') == 'on' or form_data.get('export_data') == 'true'
        import_data = form_data.get('import_data') == 'on' or form_data.get('import_data') == 'true'
        
        update_query = f'UPDATE {FASTAPI_USERS_TABLE} SET view = %s, edit = %s, add = %s, delete = %s, admin = %s, run_scripts = %s, export_data = %s, import_data = %s WHERE "user" = %s'
        cursor.execute(update_query, (view, edit, add, delete, admin, run_scripts, export_data, import_data, user_to_edit))
        conn.commit()
        cursor.close()
        conn.close()
        
        return {"success": True, "message": f"User '{user_to_edit}' permissions updated"}
    except Exception as e:
        print(f"ERROR updating user: {e}")
        return {"error": str(e)}


@app.delete("/api/admin/users/{user_to_delete}")
async def api_delete_user(user_to_delete: str, request: Request):
    """Delete a user from the system"""
    username = request.cookies.get("username")
    if not username:
        return {"error": "Not authenticated"}
    
    user_perms = get_user_permissions(username)
    if not is_user_admin(user_perms):
        return {"error": "Admin privileges required"}
    
    try:
        if delete_user(user_to_delete):
            return {"success": True, "message": f"User '{user_to_delete}' deleted successfully"}
        else:
            return {"error": f"Failed to delete user '{user_to_delete}'"}
    except Exception as e:
        print(f"ERROR in delete user endpoint: {e}")
        return {"error": str(e)}


@app.get("/api/admin/config")
async def api_get_config(request: Request):
    """Get current config.toml content"""
    username = request.cookies.get("username")
    if not username:
        return {"error": "Not authenticated"}
    
    user_perms = get_user_permissions(username)
    if not is_user_admin(user_perms):
        return {"error": "Admin privileges required"}
    
    try:
        config_path = Path(__file__).parent / "config.toml"
        with open(config_path, "r") as f:
            content = f.read()
        return {"config": content}
    except Exception as e:
        print(f"ERROR reading config: {e}")
        return {"error": str(e)}


@app.post("/api/admin/config")
async def api_update_config(request: Request):
    """Update config.toml"""
    username = request.cookies.get("username")
    if not username:
        return {"error": "Not authenticated"}
    
    user_perms = get_user_permissions(username)
    if not is_user_admin(user_perms):
        return {"error": "Admin privileges required"}
    
    try:
        body = await request.json()
        new_content = body.get("config", "")
        
        # Validate TOML syntax
        import tomllib
        try:
            tomllib.loads(new_content)
        except Exception as parse_err:
            return {"error": f"Invalid TOML syntax: {str(parse_err)}"}
        
        config_path = Path(__file__).parent / "config.toml"
        
        # Create backup
        backup_path = config_path.with_suffix(".toml.bak")
        with open(config_path, "r") as f:
            backup_content = f.read()
        with open(backup_path, "w") as f:
            f.write(backup_content)
        
        # Write new config
        with open(config_path, "w") as f:
            f.write(new_content)
        
        # Reload config cache
        from config import reload_config
        reload_config()
        
        print(f"Config updated by {username}, backup saved to {backup_path}")
        return {"success": True, "message": "Config updated successfully (backup saved)"}
    except Exception as e:
        print(f"ERROR updating config: {e}")
        return {"error": str(e)}


@app.get("/api/admin/reload-config")
async def api_reload_config(request: Request):
    """Reload configuration from disk"""
    username = request.cookies.get("username")
    if not username:
        return {"error": "Not authenticated"}
    
    user_perms = get_user_permissions(username)
    if not is_user_admin(user_perms):
        return {"error": "Admin privileges required"}
    
    try:
        from config import reload_config
        reload_config()
        print(f"Config reloaded by {username}")
        return {"success": True, "message": "Configuration reloaded from disk"}
    except Exception as e:
        print(f"ERROR reloading config: {e}")
        return {"error": f"Failed to reload config: {str(e)}"}


@app.get("/logout", response_class=HTMLResponse)
async def logout(request: Request):
    """Logout endpoint - clear session and redirect to login"""
    response = RedirectResponse(url="/", status_code=303)
    response.delete_cookie("username", path="/")
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response


@app.get("/help", response_class=HTMLResponse)
async def help(request: Request):
    """Help page"""
    username = request.cookies.get("username")
    if not username:
        return RedirectResponse(url="/", status_code=303)
    
    user_perms = get_user_permissions(username)
    
    html = render_template("help.html", {
        "username": username,
        "page": "help",
        "is_admin": is_user_admin(user_perms),
        "can_run_scripts": can_user_run_scripts(user_perms),
        "available_tables": []
    })
    return html


@app.post("/save-row", response_class=HTMLResponse)
async def save_row(request: Request, table_name: str = Form(...)):
    """Save edited row"""
    username = request.cookies.get("username")
    if not username:
        return render_template("error.html", {"error": "Not authenticated", "username": "guest"})
    
    form_data = await request.form()
    table_name = form_data.get('table_name')
    
    try:
        # Get user permissions
        user_perms = get_user_permissions(username)
        
        # Check if user can edit
        if not can_user_edit(user_perms):
            return render_template("error.html", {"error": "You do not have permission to edit data", "username": username})
        
        # Check table permissions
        perms = get_table_permissions(table_name)
        editable_columns = set(perms["editable_columns"])
        
        conn = get_db_connection()
        if not conn:
            return render_template("error.html", {"error": "Database connection failed", "username": username})
        
        columns = get_table_columns(table_name)
        col_types = get_column_types(table_name)
        id_column = columns[0]
        row_id = form_data.get(id_column)
        
        # Build UPDATE query - only include editable columns
        set_clauses = []
        values = []
        
        for col in columns[1:]:  # Skip ID
            if col not in editable_columns:
                continue  # Skip non-editable columns
            
            if col in form_data or col_types.get(col) == 'boolean':
                set_clauses.append(f"{col} = %s")
                
                # Handle booleans - checkbox sends 'on' if checked, nothing if unchecked
                if col_types.get(col) == 'boolean':
                    values.append(col in form_data and form_data.get(col) is not None)
                else:
                    values.append(form_data.get(col))
        
        if not set_clauses:
            return render_template("error.html", {"error": "No editable columns to update", "username": username})
        
        values.append(row_id)
        query = f"UPDATE {table_name} SET {', '.join(set_clauses)} WHERE {id_column} = %s"
        
        cursor = conn.cursor()
        cursor.execute(query, values)
        conn.commit()
        cursor.close()
        
        # Log audit event
        reference = f"{table_name}:id={row_id}"
        log_audit_event(username, reference, "UPDATE")
        
        return_db_connection(conn)
        
        # Redirect back to table view
        return RedirectResponse(url=f"/table/{table_name}", status_code=303)
    except Exception as e:
        print(f"ERROR saving row: {e}")
        import traceback
        traceback.print_exc()
        return render_template("error.html", {"error": str(e), "username": username})


@app.post("/add-row", response_class=HTMLResponse)
async def add_row(request: Request, table_name: str = Form(...)):
    """Add a new row"""
    username = request.cookies.get("username")
    if not username:
        return render_template("error.html", {"error": "Not authenticated", "username": "guest"})
    
    form_data = await request.form()
    table_name = form_data.get('table_name')
    
    try:
        # Get user permissions
        user_perms = get_user_permissions(username)
        
        # Check if user can add rows
        if not can_user_add(user_perms):
            return render_template("error.html", {"error": "You do not have permission to add rows", "username": username})
        
        # Check table permissions
        perms = get_table_permissions(table_name)
        if not perms["allow_add"]:
            return render_template("error.html", {"error": "Adding rows to this table is not allowed", "username": username})
        
        conn = get_db_connection()
        if not conn:
            return render_template("error.html", {"error": "Database connection failed", "username": username})
        
        columns = get_table_columns(table_name)
        col_types = get_column_types(table_name)
        
        # Skip the ID column (first column) - it auto-increments
        insert_columns = columns[1:]
        insert_values = []
        placeholders = []
        
        for col in insert_columns:
            # When adding rows, accept all columns (not just editable_columns)
            placeholders.append("%s")
            
            # Handle booleans - checkbox sends 'on' if checked, nothing if unchecked
            if col_types.get(col) == 'boolean':
                insert_values.append(col in form_data and form_data.get(col) is not None)
            else:
                insert_values.append(form_data.get(col))
        
        if not placeholders:
            return {"error": "No columns to insert"}
        
        query = f"INSERT INTO {table_name} ({', '.join(insert_columns)}) VALUES ({', '.join(placeholders)})"
        
        cursor = conn.cursor()
        cursor.execute(query, insert_values)
        conn.commit()
        
        # Get the ID of the inserted row
        id_column = columns[0]
        cursor.execute(f"SELECT lastval() as id")
        result = cursor.fetchone()
        new_row_id = result[0] if result else "unknown"
        cursor.close()
        
        # Log audit event
        reference = f"{table_name}:id={new_row_id}"
        log_audit_event(username, reference, "INSERT")
        
        return_db_connection(conn)
        
        # Redirect back to table view
        return RedirectResponse(url=f"/table/{table_name}", status_code=303)
    except Exception as e:
        print(f"ERROR adding row: {e}")
        import traceback
        traceback.print_exc()
        return render_template("error.html", {"error": str(e), "username": username})


@app.post("/delete-row", response_class=HTMLResponse)
async def delete_row(request: Request, table_name: str = Form(...)):
    """Delete a row"""
    username = request.cookies.get("username")
    if not username:
        return render_template("error.html", {"error": "Not authenticated", "username": "guest"})
    
    form_data = await request.form()
    table_name = form_data.get('table_name')
    
    try:
        # Get user permissions
        user_perms = get_user_permissions(username)
        
        # Check if user can delete rows
        if not can_user_delete(user_perms):
            return render_template("error.html", {"error": "You do not have permission to delete rows", "username": username})
        
        # Check table permissions
        perms = get_table_permissions(table_name)
        if not perms["allow_delete"]:
            return render_template("error.html", {"error": "Deleting rows from this table is not allowed", "username": username})
        
        conn = get_db_connection()
        if not conn:
            return render_template("error.html", {"error": "Database connection failed", "username": username})
        
        columns = get_table_columns(table_name)
        id_column = columns[0]
        row_id = form_data.get(id_column)
        
        query = f"DELETE FROM {table_name} WHERE {id_column} = %s"
        
        cursor = conn.cursor()
        cursor.execute(query, (row_id,))
        conn.commit()
        cursor.close()
        
        # Log audit event
        reference = f"{table_name}:id={row_id}"
        log_audit_event(username, reference, "DELETE")
        
        return_db_connection(conn)
        
        # Redirect back to table view
        return RedirectResponse(url=f"/table/{table_name}", status_code=303)
    except Exception as e:
        print(f"ERROR deleting row: {e}")
        import traceback
        traceback.print_exc()
        return render_template("error.html", {"error": str(e), "username": username})


@app.put("/api/table/{table_name}/{row_id}")
async def api_update_row(table_name: str, row_id: str, request: Request):
    """API endpoint to update a row in the table"""
    username = request.cookies.get("username")
    if not username:
        return {"error": "Not authenticated"}
    
    try:
        data = await request.json()
        conn = get_db_connection()
        if not conn:
            return {"error": "Database connection failed"}
        
        # Get column names
        columns = get_table_columns(table_name)
        if not columns:
            return {"error": "Table not found"}
        
        # Build UPDATE query
        id_column = columns[0]  # Assume first column is primary key
        set_clauses = []
        values = []
        
        for col in columns[1:]:  # Skip ID column
            if col in data:
                set_clauses.append(f"{col} = %s")
                values.append(data[col])
        
        if not set_clauses:
            return {"error": "No columns to update"}
        
        values.append(row_id)  # Add ID value for WHERE clause
        
        query = f"UPDATE {table_name} SET {', '.join(set_clauses)} WHERE {id_column} = %s"
        
        cursor = conn.cursor()
        cursor.execute(query, values)
        conn.commit()
        cursor.close()
        conn.close()
        
        return {"success": True, "message": "Row updated"}
    except Exception as e:
        print(f"ERROR updating row: {e}")
        import traceback
        traceback.print_exc()
        return {"error": str(e)}


@app.get("/api/tables")
async def api_get_tables(request: Request):
    """API endpoint to get available tables with display names (async)"""
    username = request.cookies.get("username")
    if not username:
        return {"tables": []}
    
    tables = get_tables_with_display_names()
    return {"tables": tables}


@app.get("/api/disk-stats")
async def api_get_disk_stats(request: Request):
    """API endpoint to get disk usage statistics"""
    username = request.cookies.get("username")
    if not username:
        return {"error": "unauthorized"}
    
    stats = get_disk_stats()
    return stats


@app.get("/table/{table_name}", response_class=HTMLResponse)
async def get_table_data_route(table_name: str, request: Request, page: int = 1, limit: int = 100):
    """Get table data for editing with pagination"""
    username = request.cookies.get("username")
    if not username:
        return RedirectResponse(url="/", status_code=303)
    
    # Get user permissions and check if they can view
    user_perms = get_user_permissions(username)
    if not can_user_view(user_perms):
        return render_template("error.html", {"error": "You do not have permission to view data", "username": username})
    
    # Validate page number
    page = max(1, page)
    offset = (page - 1) * limit
    
    columns = get_table_columns(table_name)
    col_types = get_column_types(table_name)
    data, total_rows = get_table_data(table_name, limit=limit, offset=offset)
    
    # Calculate pagination info
    total_pages = (total_rows + limit - 1) // limit  # Ceiling division
    has_prev = page > 1
    has_next = page < total_pages
    
    # Get table permissions
    perms = get_table_permissions(table_name)
    
    # Get lookup configuration for this table
    lookups_config = config.get("lookups", {})
    lookup_cols = lookups_config.get(table_name, [])
    
    # Build lookups dict for template (col -> table.col)
    lookups = {col: f"{table_name}.{col}" for col in lookup_cols}
    
    # Pre-fetch all lookup options
    lookup_options = {}
    for col in lookup_cols:
        source_def = lookups[col]  # Format: "table.column"
        options = get_lookup_options(source_def, col)
        lookup_options[col] = options
    
    # Get age coloring configuration
    age_coloring_config = config.get("age_coloring", {})
    age_coloring_enabled = age_coloring_config.get("enabled", False)
    age_coloring_column = None
    
    if age_coloring_enabled:
        # Check if this table is in the list of tables to apply age coloring
        age_coloring_tables = age_coloring_config.get("tables", [])
        should_apply = not age_coloring_tables or table_name in age_coloring_tables
        
        if should_apply:
            # Find first matching column in this table
            possible_columns = age_coloring_config.get("columns", ["last_updated", "created_date", "updated_at"])
            for col in possible_columns:
                if col in columns:
                    age_coloring_column = col
                    break
    
    html = render_template("table.html", {
        "columns": columns,
        "col_types": col_types,
        "data": data,
        "table_name": table_name,
        "username": username,
        "page": "table",
        "is_admin": is_user_admin(user_perms),
        "can_run_scripts": can_user_run_scripts(user_perms),
        "available_tables": [],
        "lookups": lookups,
        "lookup_options": lookup_options,
        "allow_add": perms["allow_add"] and can_user_add(user_perms),
        "allow_delete": perms["allow_delete"] and can_user_delete(user_perms),
        "editable_columns": perms["editable_columns"] if can_user_edit(user_perms) else [],
        # Pagination info
        "current_page": page,
        "total_rows": total_rows,
        "total_pages": total_pages,
        "page_size": limit,
        "has_prev": has_prev,
        "has_next": has_next,
        "prev_page": page - 1,
        "next_page": page + 1,
        # Export/Import permissions
        "can_export": user_perms.get('export_data', False),
        "can_import": user_perms.get('import_data', False),
    })
    return html


# ============================================================================
# LOOKUP API ENDPOINTS
# ============================================================================

@app.get("/api/lookup-options/{source_table}/{column_name}")
async def api_get_lookup_options(source_table: str, column_name: str, request: Request):
    """Get dropdown options for a lookup field"""
    username = request.cookies.get("username")
    if not username:
        return {"error": "Not authenticated"}
    
    try:
        # Format: "i07_gsm.gsm_name"
        lookup_key = f"{source_table}.{column_name}"
        options = get_lookup_options(lookup_key, column_name)
        return {"options": options}
    except Exception as e:
        print(f"ERROR fetching lookup options: {e}")
        import traceback
        traceback.print_exc()
        return {"error": str(e)}


@app.post("/api/add-lookup-value")
async def api_add_lookup_value(request: Request):
    """Add a new value to a lookup source table"""
    username = request.cookies.get("username")
    if not username:
        return {"error": "Not authenticated"}
    
    try:
        data = await request.json()
        source_table = data.get("source_table")  # e.g., "i07_gsm"
        column_name = data.get("column_name")    # e.g., "gsm_name"
        value = data.get("value")                 # e.g., "gsm_new"
        
        if not all([source_table, column_name, value]):
            return {"error": "Missing required fields: source_table, column_name, value"}
        
        # Format: "table.column"
        lookup_key = f"{source_table}.{column_name}"
        success = add_lookup_value(lookup_key, column_name, value)
        
        if success:
            return {"success": True, "message": f"Added {value} to {lookup_key}"}
        else:
            return {"error": "Failed to add lookup value"}
    except Exception as e:
        print(f"ERROR adding lookup value: {e}")
        import traceback
        traceback.print_exc()
        return {"error": str(e)}


@app.get("/api/field-lookups/{table_name}")
async def api_get_field_lookups(table_name: str, request: Request):
    """Get lookup configuration for a table"""
    username = request.cookies.get("username")
    if not username:
        return {"error": "Not authenticated"}
    
    try:
        # Get lookups from config for this table
        lookups_config = config.get("lookups", {})
        lookup_cols = lookups_config.get(table_name, [])
        lookups = {col: f"{table_name}.{col}" for col in lookup_cols}
        return {"lookups": lookups}
    except Exception as e:
        print(f"ERROR fetching field lookups: {e}")
        return {"error": str(e)}


@app.post("/api/generate-lookups/{table_name}")
async def api_generate_lookups(table_name: str, request: Request):
    """Generate lookups for a table by scanning TEXT columns"""
    username = request.cookies.get("username")
    if not username:
        return {"error": "Not authenticated"}
    
    try:
        # Generate lookups
        lookups = generate_lookups_for_table(table_name)
        
        if not lookups:
            return {"success": False, "message": f"No TEXT columns found in {table_name}"}
        
        # Update config[lookups]
        if "lookups" not in config:
            config["lookups"] = {}
        
        # Add or update table in lookups
        existing = config["lookups"].get(table_name, [])
        new_cols = [col for col in lookups.keys() if col not in existing]
        
        if not new_cols:
            return {"success": False, "message": f"All columns already in config for {table_name}"}
        
        config["lookups"][table_name] = list(lookups.keys())
        
        # Write config back to file
        from pathlib import Path
        config_path = Path(__file__).parent / "config.toml"
        with open(config_path, 'w') as f:
            # Write sections in order
            for section in ['database', 'tables', 'dashboard', 'lookups']:
                if section in config:
                    f.write(f"[{section}]\n")
                    section_data = config[section]
                    
                    if section == 'lookups':
                        # Format: table = ["col1", "col2"]
                        for tbl, cols in sorted(section_data.items()):
                            cols_str = ', '.join(f'"{col}"' for col in cols)
                            f.write(f'{tbl} = [{cols_str}]\n')
                    else:
                        for key, value in section_data.items():
                            if isinstance(value, list):
                                if all(isinstance(v, str) for v in value):
                                    items = ', '.join(f'"{v}"' for v in value)
                                    f.write(f'{key} = [{items}]\n')
                                else:
                                    f.write(f'{key} = {value}\n')
                            elif isinstance(value, str):
                                f.write(f'{key} = "{value}"\n')
                            else:
                                f.write(f'{key} = {value}\n')
                    f.write('\n')
        
        return {
            "success": True,
            "message": f"Generated lookups for {table_name}",
            "added_columns": new_cols,
            "all_lookups": list(lookups.keys())
        }
    except Exception as e:
        print(f"ERROR generating lookups: {e}")
        import traceback
        traceback.print_exc()
        return {"success": False, "error": str(e)}


# ============================================================================
# SHELL SCRIPT EXECUTION ENDPOINTS
# ============================================================================

@app.get("/api/scripts")
async def api_get_scripts(request: Request):
    """Get list of available shell scripts"""
    username = request.cookies.get("username")
    if not username:
        return {"error": "Not authenticated"}
    
    user_perms = get_user_permissions(username)
    if not can_user_run_scripts(user_perms):
        return {"error": "You do not have permission to run scripts"}
    
    try:
        scripts_list = []
        for script_name, script_path in SCRIPTS_CONFIG.items():
            # Skip the run_in_background subsection
            if script_name != "run_in_background" and isinstance(script_path, str):
                scripts_list.append({"name": script_name, "path": script_path})
        return {"scripts": scripts_list}
    except Exception as e:
        print(f"ERROR fetching scripts: {e}")
        return {"error": str(e)}


def execute_script_internal(script_name: str):
    """Internal function to execute a script and return result dict"""
    start_time = time.time()
    
    try:
        # Get script path from config
        if script_name not in SCRIPTS_CONFIG or SCRIPTS_CONFIG[script_name] == "run_in_background":
            return {"error": f"Script '{script_name}' not found in configuration", "success": False}
        
        script_path = SCRIPTS_CONFIG[script_name]
        if not isinstance(script_path, str):
            return {"error": f"Script '{script_name}' path is invalid", "success": False}
        
        # Resolve path relative to app.py location
        app_dir = Path(__file__).parent
        full_script_path = (app_dir / script_path).resolve()
        
        # Security check: ensure script path exists and is readable
        if not full_script_path.exists():
            return {"error": f"Script file not found: {full_script_path}", "success": False}
        
        if not full_script_path.is_file():
            return {"error": f"Script path is not a file: {full_script_path}", "success": False}
        
        # Execute the script
        print(f"Executing script: {full_script_path}")
        result = subprocess.run(
            [str(full_script_path)],
            capture_output=True,
            text=True,
            timeout=300  # 5 minute timeout
        )
        
        elapsed_time = time.time() - start_time
        
        return {
            "success": result.returncode == 0,
            "script_name": script_name,
            "return_code": result.returncode,
            "stdout": result.stdout,
            "stderr": result.stderr,
            "execution_time": elapsed_time,
        }
    except subprocess.TimeoutExpired:
        elapsed_time = time.time() - start_time
        return {"error": "Script execution timed out after 5 minutes", "success": False, "execution_time": elapsed_time}
    except Exception as e:
        elapsed_time = time.time() - start_time
        print(f"ERROR executing script: {e}")
        import traceback
        traceback.print_exc()
        return {"error": str(e), "success": False, "execution_time": elapsed_time}


def execute_script_async(script_name: str, execution_id: str):
    """Execute script in a background thread"""
    result = execute_script_internal(script_name)
    update_job_result(execution_id, result)


@app.get("/execute-script/{script_name}", response_class=HTMLResponse)
async def execute_script(script_name: str, request: Request):
    """Execute a predefined shell script and show results page"""
    username = request.cookies.get("username")
    if not username:
        return RedirectResponse(url="/", status_code=303)
    
    user_perms = get_user_permissions(username)
    if not can_user_run_scripts(user_perms):
        return render_template("error.html", {"error": "You do not have permission to run scripts", "username": username})
    
    # Check if script should run in background
    if is_script_background(script_name):
        # Execute in background
        execution_id = generate_execution_id()
        add_job(execution_id, script_name)
        
        # Start execution in background thread
        thread = threading.Thread(target=execute_script_async, args=(script_name, execution_id), daemon=True)
        thread.start()
        
        # Return a page that will poll for status
        html = render_template("script_results.html", {
            "username": username,
            "is_admin": is_user_admin(user_perms),
            "can_run_scripts": can_user_run_scripts(user_perms),
            "script_name": script_name,
            "execution_id": execution_id,
            "is_background": True,
        })
        return html
    else:
        # Execute synchronously (old behavior)
        result = execute_script_internal(script_name)
        
        html = render_template("script_results.html", {
            "username": username,
            "is_admin": is_user_admin(user_perms),
            "can_run_scripts": can_user_run_scripts(user_perms),
            "success": result.get("success", False),
            "script_name": result.get("script_name", script_name),
            "return_code": result.get("return_code", ""),
            "stdout": result.get("stdout", ""),
            "stderr": result.get("stderr", ""),
            "error": result.get("error", ""),
            "execution_time": result.get("execution_time", 0),
        })
        return html


@app.get("/script-results/{execution_id}", response_class=HTMLResponse)
async def view_script_results(execution_id: str, request: Request):
    """View results of a specific background script execution"""
    username = request.cookies.get("username")
    if not username:
        return RedirectResponse(url="/", status_code=303)
    
    user_perms = get_user_permissions(username)
    if not can_user_run_scripts(user_perms):
        return render_template("error.html", {"error": "You do not have permission to view scripts", "username": username})
    
    # Get the job info
    with _jobs_lock:
        job_info = _script_jobs.get(execution_id, {})
    
    if not job_info:
        return render_template("error.html", {"error": "Execution not found", "username": username})
    
    script_name = job_info.get("script_name", "Unknown")
    status = get_job_status(execution_id)
    
    # Return the script results page with polling if still running
    html = render_template("script_results.html", {
        "username": username,
        "is_admin": is_user_admin(user_perms),
        "can_run_scripts": can_user_run_scripts(user_perms),
        "script_name": script_name,
        "execution_id": execution_id,
        "is_background": True,
    })
    return html


@app.post("/api/execute-script/{script_name}")
async def api_execute_script(script_name: str, request: Request):
    """API endpoint to execute a script (JSON response)"""
    username = request.cookies.get("username")
    if not username:
        return {"error": "Not authenticated"}
    
    user_perms = get_user_permissions(username)
    if not can_user_run_scripts(user_perms):
        return {"error": "You do not have permission to run scripts"}
    
    # Check if script should run in background
    if is_script_background(script_name):
        # Execute in background
        execution_id = generate_execution_id()
        add_job(execution_id, script_name)
        
        # Start execution in background thread
        thread = threading.Thread(target=execute_script_async, args=(script_name, execution_id), daemon=True)
        thread.start()
        
        return {
            "success": True,
            "execution_id": execution_id,
            "message": f"Script '{script_name}' started in background",
            "is_background": True,
        }
    else:
        # Execute synchronously (old behavior)
        result = execute_script_internal(script_name)
        result["message"] = "Script executed successfully" if result.get("success") else "Script completed with errors"
        return result


@app.get("/api/script-status/{execution_id}")
async def api_script_status(execution_id: str, request: Request):
    """API endpoint to get the status of a background script execution"""
    username = request.cookies.get("username")
    if not username:
        return {"error": "Not authenticated"}
    
    user_perms = get_user_permissions(username)
    if not can_user_run_scripts(user_perms):
        return {"error": "You do not have permission to run scripts"}
    
    return get_job_status(execution_id)


@app.get("/api/active-scripts")
async def api_active_scripts(request: Request):
    """API endpoint to get list of currently running scripts"""
    username = request.cookies.get("username")
    if not username:
        return {"error": "Not authenticated"}
    
    user_perms = get_user_permissions(username)
    if not can_user_run_scripts(user_perms):
        return {"error": "You do not have permission to run scripts"}
    
    active = get_active_scripts()
    return {"scripts": active, "count": len(active)}


@app.get("/background-scripts", response_class=HTMLResponse)
async def background_scripts_page(request: Request):
    """Page to view all active and recently completed background scripts"""
    username = request.cookies.get("username")
    if not username:
        return RedirectResponse(url="/login", status_code=302)
    
    user_perms = get_user_permissions(username)
    if not can_user_run_scripts(user_perms):
        return RedirectResponse(url="/dashboard", status_code=302)
    
    # Get all jobs (active and completed)
    with _jobs_lock:
        all_jobs = list(_script_jobs.items())
    
    # Convert to display format with status info
    jobs_display = []
    for execution_id, job_info in all_jobs:
        status = get_job_status(execution_id)
        jobs_display.append({
            "execution_id": execution_id,
            "script_name": job_info.get("script_name", "Unknown"),
            "status": status.get("status", "unknown"),
            "completion_time": status.get("completion_time", None)
        })
    
    # Sort by completion time (newest first), with active jobs at top
    jobs_display.sort(key=lambda x: (x["status"] != "completed", x.get("completion_time", 0)), reverse=True)
    
    return render_template(
        "background_scripts.html",
        {
            "jobs": jobs_display,
            "page": "background-scripts",
            "username": username,
            "is_admin": user_perms.get("is_admin", False),
            "can_run_scripts": can_user_run_scripts(user_perms),
            "kerberos_domain": config.get("app", {}).get("kerberos_domain", "LOCAL")
        }
    )


# ============================================================================
# IMPORT/EXPORT ENDPOINTS
# ============================================================================

@app.get("/export-table/{table_name}")
async def export_table(table_name: str, request: Request, format: str = "csv"):
    """Export table data as CSV or Excel"""
    username = request.cookies.get("username")
    if not username:
        return JSONResponse({"error": "Not authenticated"}, status_code=401)
    
    # Check permissions
    user_perms = get_user_permissions(username)
    if not user_perms.get('view', False):
        return JSONResponse({"error": "You do not have permission to view tables"}, status_code=403)
    
    if not user_perms.get('export_data', False):
        return JSONResponse({"error": "You do not have permission to export data"}, status_code=403)
    
    # Validate table name
    available_tables = get_available_tables()
    if table_name not in available_tables:
        return JSONResponse({"error": f"Table '{table_name}' not found"}, status_code=404)
    
    try:
        if format.lower() == "csv":
            data = export_table_to_csv(table_name)
            return StreamingResponse(
                iter([data]),
                media_type="text/csv",
                headers={"Content-Disposition": f"attachment; filename={table_name}.csv"}
            )
        elif format.lower() == "xlsx":
            data = export_table_to_excel(table_name)
            return StreamingResponse(
                iter([data]),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={table_name}.xlsx"}
            )
        else:
            return JSONResponse({"error": f"Unsupported format: {format}"}, status_code=400)
    
    except Exception as e:
        print(f"ERROR exporting table {table_name}: {e}")
        import traceback
        traceback.print_exc()
        return JSONResponse({"error": f"Export failed: {str(e)}"}, status_code=500)


@app.post("/api/import-preview/{table_name}")
async def import_preview(table_name: str, request: Request):
    """Preview imported file (show first 5 rows, validate, check for duplicates)"""
    from database import validate_import_row, get_table_columns
    import csv
    import io
    
    username = request.cookies.get("username")
    if not username:
        return {"error": "Not authenticated", "status": "error"}
    
    # Check permissions
    user_perms = get_user_permissions(username)
    if not user_perms.get('import_data', False):
        return {"error": "You do not have permission to import data", "status": "error"}
    
    # Validate table name
    available_tables = get_available_tables()
    if table_name not in available_tables:
        return {"error": f"Table '{table_name}' not found", "status": "error"}
    
    try:
        # Get uploaded file
        form = await request.form()
        file = form.get("file")
        
        if not file:
            return {"error": "No file uploaded", "status": "error"}
        
        # Read file content
        content = await file.read()
        file_type = file.filename.split('.')[-1].lower()
        
        rows = []
        columns = []
        
        if file_type == 'csv':
            # Parse CSV
            content_str = content.decode('utf-8')
            reader = csv.DictReader(io.StringIO(content_str))
            columns = reader.fieldnames or []
            rows = list(reader)
        
        elif file_type == 'xlsx':
            # Parse Excel - with fallback for style parsing errors
            try:
                from openpyxl import load_workbook
                wb = load_workbook(io.BytesIO(content), data_only=True)
                ws = wb.active
                
                # Read header - get all values from first row
                columns = []
                for cell in ws[1]:
                    if cell.value is not None:
                        columns.append(str(cell.value))
                    else:
                        break  # Stop at first empty cell
                
                # Read rows
                for row_idx in range(2, ws.max_row + 1):
                    row_dict = {}
                    has_data = False
                    for col_idx, col_name in enumerate(columns, 1):
                        cell_value = ws.cell(row=row_idx, column=col_idx).value
                        row_dict[col_name] = cell_value
                        if cell_value is not None:
                            has_data = True
                    if has_data:  # Only add rows with data
                        rows.append(row_dict)
            except Exception as excel_error:
                # Fallback: parse Excel as ZIP to extract data without loading styles
                import zipfile
                import xml.etree.ElementTree as ET
                try:
                    with zipfile.ZipFile(io.BytesIO(content)) as zf:
                        # Read shared strings first
                        shared_strings = []
                        try:
                            strings_xml = zf.read('xl/sharedStrings.xml')
                            strings_root = ET.fromstring(strings_xml)
                            for si_elem in strings_root.findall('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}si'):
                                t_elem = si_elem.find('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t')
                                if t_elem is not None and t_elem.text:
                                    shared_strings.append(t_elem.text)
                        except:
                            pass  # No shared strings or error reading them
                        
                        # Read the worksheet XML
                        sheet_names = [n for n in zf.namelist() if n.startswith('xl/worksheets/sheet')]
                        if not sheet_names:
                            return {"error": "No worksheets found in Excel file", "status": "error"}
                        
                        sheet_xml = zf.read(sheet_names[0])
                        root = ET.fromstring(sheet_xml)
                        
                        # Parse rows
                        rows_elem = root.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}row')
                        columns = []
                        
                        for row_idx, row_elem in enumerate(rows_elem):
                            row_dict = {}
                            cells = row_elem.findall('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}c')
                            
                            for col_idx, cell_elem in enumerate(cells):
                                # Get cell type
                                cell_type = cell_elem.get('t', 'n')  # Default to number type
                                
                                # Get cell value
                                value_elem = cell_elem.find('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v')
                                value = value_elem.text if value_elem is not None else None
                                
                                # If type is 's', it's a string index - look it up in shared strings
                                if cell_type == 's' and value is not None:
                                    try:
                                        str_idx = int(value)
                                        if str_idx < len(shared_strings):
                                            value = shared_strings[str_idx]
                                    except:
                                        pass
                                
                                if row_idx == 0:  # Header row
                                    if value:
                                        columns.append(str(value))
                                else:  # Data rows
                                    if col_idx < len(columns):
                                        row_dict[columns[col_idx]] = value
                            
                            if row_idx > 0 and row_dict:  # Skip header, skip empty rows
                                rows.append(row_dict)
                except Exception as fallback_error:
                    return {"error": f"Failed to parse Excel file: {str(fallback_error)}", "status": "error"}
        
        else:
            return {"error": f"Unsupported file format: {file_type}", "status": "error"}
        
        # Preview: show first 20 rows + validation
        preview_rows = []
        validation_errors = []
        validation_warnings = []
        duplicate_count = 0
        
        for row_idx, row in enumerate(rows[:20], 1):
            is_valid, errors, warnings = validate_import_row(table_name, row)
            preview_rows.append({
                'data': row,
                'valid': is_valid,
                'errors': errors,
                'warnings': warnings
            })
            validation_errors.extend([f"Row {row_idx}: {e}" for e in errors])
            validation_warnings.extend([f"Row {row_idx}: {w}" for w in warnings])
        
        # Check for total rows and duplicates
        total_rows = len(rows)
        
        return {
            "status": "ok",
            "filename": file.filename,
            "file_type": file_type,
            "columns": columns,
            "total_rows": total_rows,
            "preview_rows": preview_rows,
            "validation_errors": validation_errors,
            "validation_warnings": validation_warnings,
            "message": f"Preview loaded: {total_rows} rows, {len(columns)} columns"
        }
    
    except Exception as e:
        print(f"ERROR in import preview: {e}")
        import traceback
        traceback.print_exc()
        return {"error": f"Preview failed: {str(e)}", "status": "error"}


@app.post("/api/import-execute/{table_name}")
async def import_execute(table_name: str, request: Request):
    """Execute import (insert rows with validation and duplicate handling)"""
    from database import insert_imported_rows
    import csv
    import io
    
    username = request.cookies.get("username")
    if not username:
        return {"error": "Not authenticated", "status": "error"}
    
    # Check permissions
    user_perms = get_user_permissions(username)
    if not user_perms.get('import_data', False):
        return {"error": "You do not have permission to import data", "status": "error"}
    
    # Validate table name
    available_tables = get_available_tables()
    if table_name not in available_tables:
        return {"error": f"Table '{table_name}' not found", "status": "error"}
    
    try:
        # Get form data
        form = await request.form()
        file = form.get("file")
        replace_all = form.get("replace_all", "false").lower() == "true"
        skip_duplicates = form.get("skip_duplicates", "true").lower() == "true"
        dry_run = form.get("dry_run", "false").lower() == "true"
        
        if not file:
            return {"error": "No file uploaded", "status": "error"}
        
        # Read file content
        content = await file.read()
        file_type = file.filename.split('.')[-1].lower()
        
        rows = []
        
        if file_type == 'csv':
            # Parse CSV
            content_str = content.decode('utf-8')
            reader = csv.DictReader(io.StringIO(content_str))
            rows = list(reader)
        
        elif file_type == 'xlsx':
            # Parse Excel - with fallback for style parsing errors
            try:
                from openpyxl import load_workbook
                wb = load_workbook(io.BytesIO(content), data_only=True)
                ws = wb.active
                
                # Read header - get all values from first row
                columns = []
                for cell in ws[1]:
                    if cell.value is not None:
                        columns.append(str(cell.value))
                    else:
                        break  # Stop at first empty cell
                
                # Read rows
                for row_idx in range(2, ws.max_row + 1):
                    row_dict = {}
                    has_data = False
                    for col_idx, col_name in enumerate(columns, 1):
                        cell_value = ws.cell(row=row_idx, column=col_idx).value
                        row_dict[col_name] = cell_value
                        if cell_value is not None:
                            has_data = True
                    if has_data:  # Only add rows with data
                        rows.append(row_dict)
            except Exception as excel_error:
                # Fallback: parse Excel as ZIP to extract data without loading styles
                import zipfile
                import xml.etree.ElementTree as ET
                try:
                    with zipfile.ZipFile(io.BytesIO(content)) as zf:
                        # Read shared strings first
                        shared_strings = []
                        try:
                            strings_xml = zf.read('xl/sharedStrings.xml')
                            strings_root = ET.fromstring(strings_xml)
                            for si_elem in strings_root.findall('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}si'):
                                t_elem = si_elem.find('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t')
                                if t_elem is not None and t_elem.text:
                                    shared_strings.append(t_elem.text)
                        except:
                            pass  # No shared strings or error reading them
                        
                        # Read the worksheet XML
                        sheet_names = [n for n in zf.namelist() if n.startswith('xl/worksheets/sheet')]
                        if not sheet_names:
                            return {"error": "No worksheets found in Excel file", "status": "error"}
                        
                        sheet_xml = zf.read(sheet_names[0])
                        root = ET.fromstring(sheet_xml)
                        
                        # Parse rows
                        rows_elem = root.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}row')
                        columns = []
                        
                        for row_idx, row_elem in enumerate(rows_elem):
                            row_dict = {}
                            cells = row_elem.findall('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}c')
                            
                            for col_idx, cell_elem in enumerate(cells):
                                # Get cell type
                                cell_type = cell_elem.get('t', 'n')  # Default to number type
                                
                                # Get cell value
                                value_elem = cell_elem.find('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v')
                                value = value_elem.text if value_elem is not None else None
                                
                                # If type is 's', it's a string index - look it up in shared strings
                                if cell_type == 's' and value is not None:
                                    try:
                                        str_idx = int(value)
                                        if str_idx < len(shared_strings):
                                            value = shared_strings[str_idx]
                                    except:
                                        pass
                                
                                if row_idx == 0:  # Header row
                                    if value:
                                        columns.append(str(value))
                                else:  # Data rows
                                    if col_idx < len(columns):
                                        row_dict[columns[col_idx]] = value
                            
                            if row_idx > 0 and row_dict:  # Skip header, skip empty rows
                                rows.append(row_dict)
                except Exception as fallback_error:
                    return {"error": f"Failed to parse Excel file: {str(fallback_error)}", "status": "error"}
        
        else:
            return {"error": f"Unsupported file format: {file_type}", "status": "error"}
        
        # Execute import
        result = insert_imported_rows(table_name, rows, skip_duplicates=skip_duplicates, dry_run=dry_run, replace_all=replace_all)
        
        # Format response
        message_parts = [f"Inserted: {result['inserted']}"]
        if skip_duplicates and result['skipped'] > 0:
            message_parts.append(f"Skipped: {result['skipped']} (duplicates)")
        
        if dry_run:
            message_parts.append("(DRY-RUN MODE - data not saved)")
        
        return {
            "status": "ok",
            "message": ", ".join(message_parts),
            "inserted": result['inserted'],
            "skipped": result['skipped'],
            "errors": result['errors'],
            "warnings": result['warnings'],
            "dry_run": dry_run
        }
    
    except Exception as e:
        print(f"ERROR in import execute: {e}")
        import traceback
        traceback.print_exc()
        return {"error": f"Import failed: {str(e)}", "status": "error"}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)

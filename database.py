"""Database operations"""
import psycopg2
from psycopg2 import pool
from psycopg2.extras import RealDictCursor
import re
from config import DB_CONFIG, config

# ============================================================================
# CONNECTION POOLING
# ============================================================================

_db_pool = None

def initialize_connection_pool():
    """Initialize database connection pool on startup"""
    global _db_pool
    try:
        _db_pool = psycopg2.pool.ThreadedConnectionPool(
            minconn=2,
            maxconn=10,
            **DB_CONFIG
        )
        print("✓ Database connection pool initialized")
    except Exception as e:
        print(f"ERROR initializing connection pool: {e}")
        import traceback
        traceback.print_exc()
        raise


def get_db_connection():
    """Get PostgreSQL connection from pool"""
    global _db_pool
    if _db_pool is None:
        initialize_connection_pool()
    
    try:
        return _db_pool.getconn()
    except Exception as e:
        print(f"DB CONNECTION ERROR: {e}")
        import traceback
        traceback.print_exc()
        return None


def return_db_connection(conn):
    """Return connection to pool"""
    global _db_pool
    if _db_pool is not None and conn is not None:
        _db_pool.putconn(conn)


def filter_tables(tables: list) -> list:
    """Filter tables based on include/exclude patterns from config"""
    include_patterns = config.get("tables", {}).get("include", [])
    exclude_patterns = config.get("tables", {}).get("exclude", [])
    
    if not include_patterns:
        include_patterns = [".*"]  # Include all by default if not specified
    
    filtered = []
    
    for table in tables:
        # Check if table matches any exclude pattern
        excluded = False
        for pattern in exclude_patterns:
            if re.match(pattern, table):
                excluded = True
                break
        
        if excluded:
            continue
        
        # Check if table matches any include pattern
        included = False
        for pattern in include_patterns:
            if re.match(pattern, table):
                included = True
                break
        
        if included:
            filtered.append(table)
    
    return filtered


def get_available_tables():
    """Get list of available tables in public schema (filtered by config)"""
    conn = get_db_connection()
    if not conn:
        return []
    
    try:
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        # Use pg_tables which is more reliable for getting user tables
        cursor.execute(
            "SELECT tablename FROM pg_tables WHERE schemaname = 'public' ORDER BY tablename"
        )
        tables = [row['tablename'] for row in cursor.fetchall()]
        cursor.close()
        
        # Apply filtering from config
        filtered_tables = filter_tables(tables)
        return filtered_tables
    except Exception as e:
        print(f"ERROR fetching tables: {e}")
        import traceback
        traceback.print_exc()
        return []
    finally:
        return_db_connection(conn)


def get_table_columns(table_name: str):
    """Get columns from specified table"""
    conn = get_db_connection()
    if not conn:
        print(f"ERROR: Could not connect to database for table '{table_name}'")
        return []
    
    try:
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        query = "SELECT column_name FROM information_schema.columns WHERE table_name = %s ORDER BY ordinal_position"
        print(f"DEBUG: Executing query: {query} with table_name='{table_name}'")
        cursor.execute(query, (table_name,))
        columns = [row['column_name'] for row in cursor.fetchall()]
        print(f"Columns found for table '{table_name}': {columns}")
        cursor.close()
        return columns
    except Exception as e:
        print(f"ERROR fetching columns for '{table_name}': {e}")
        import traceback
        traceback.print_exc()
        return []
    finally:
        return_db_connection(conn)


def get_column_types(table_name: str):
    """Get column names and their data types"""
    conn = get_db_connection()
    if not conn:
        print("ERROR: Could not connect to database")
        return {}
    
    try:
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        cursor.execute(
            "SELECT column_name, data_type FROM information_schema.columns WHERE table_name = %s ORDER BY ordinal_position",
            (table_name,)
        )
        col_types = {row['column_name']: row['data_type'] for row in cursor.fetchall()}
        print(f"Column types for table '{table_name}': {col_types}")
        cursor.close()
        return col_types
    except Exception as e:
        print(f"ERROR fetching column types: {e}")
        import traceback
        traceback.print_exc()
        return {}
    finally:
        return_db_connection(conn)


def get_table_data(table_name: str, limit: int = 100, offset: int = 0):
    """Get data from specified table with pagination"""
    conn = get_db_connection()
    if not conn:
        print("ERROR: Could not connect to database")
        return [], 0
    
    try:
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        # Get column names to sort by first column (usually the primary key)
        cursor.execute(f"SELECT column_name FROM information_schema.columns WHERE table_name = %s ORDER BY ordinal_position LIMIT 1", (table_name,))
        first_col_result = cursor.fetchone()
        first_col = first_col_result['column_name'] if first_col_result else 'id'
        
        # Get total count
        count_query = f"SELECT COUNT(*) as count FROM {table_name}"
        cursor.execute(count_query)
        count_result = cursor.fetchone()
        total_rows = count_result['count'] if count_result else 0
        
        query = f"SELECT * FROM {table_name} ORDER BY {first_col} LIMIT %s OFFSET %s"
        print(f"DEBUG: Executing query: {query} with LIMIT {limit} OFFSET {offset}")
        cursor.execute(query, (limit, offset))
        data = cursor.fetchall()
        print(f"Rows fetched from '{table_name}': {len(data)} (total: {total_rows})")
        cursor.close()
        return data, total_rows
    except Exception as e:
        print(f"ERROR fetching data from '{table_name}': {e}")
        import traceback
        traceback.print_exc()
        return [], 0
    finally:
        return_db_connection(conn)


# ============================================================================
# LOOKUP FUNCTIONS
# ============================================================================

def get_text_columns(table_name: str) -> list:
    """Get all TEXT columns from a table"""
    col_types = get_column_types(table_name)
    text_cols = [col for col, dtype in col_types.items() if dtype == 'text']
    return text_cols


def find_source_table(column_name: str) -> str:
    """
    Find source table for a column by naming convention.
    Examples:
    - gsm_name -> i07_gsm.gsm_name
    - wsl_name -> i07_wsl.wsl_name
    """
    available_tables = get_available_tables()
    
    # Extract base name from column (e.g., "gsm_name" -> "gsm")
    base_name = column_name.rsplit('_', 1)[0]
    
    # Look for matching table
    for table in available_tables:
        # Match i07_gsm, i07_wsl, etc.
        if base_name in table.lower() or table.lower().endswith(base_name):
            # Verify column exists in potential source table
            source_cols = get_table_columns(table)
            if column_name in source_cols:
                return f"{table}.{column_name}"
            # Also try just the base column name
            if base_name in source_cols:
                return f"{table}.{base_name}"
    
    return None


def get_lookup_options(source_table: str, column_name: str) -> list:
    """
    Fetch distinct values from a source table column for dropdown.
    Format: "table.column"
    Returns list of tuples: [(value, value), ...] for simple values
    """
    conn = get_db_connection()
    if not conn:
        return []
    
    try:
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        table, col = source_table.split('.')
        
        # Get distinct values, ordered case-insensitively using subquery
        # PostgreSQL doesn't allow ORDER BY expressions not in SELECT with DISTINCT
        query = f"SELECT {col} FROM (SELECT DISTINCT {col} FROM {table}) sub ORDER BY LOWER({col}), {col}"
        print(f"DEBUG: Lookup query: {query}")
        cursor.execute(query)
        results = cursor.fetchall()
        
        # Return list of values
        values = [row[col] for row in results]
        cursor.close()
        return values
    except Exception as e:
        print(f"ERROR fetching lookup options from {source_table}: {e}")
        import traceback
        traceback.print_exc()
        return []
    finally:
        return_db_connection(conn)


def add_lookup_value(source_table: str, column_name: str, value: str) -> bool:
    """
    Add a new value to a lookup source table.
    Format: source_table = "table.column"
    """
    conn = get_db_connection()
    if not conn:
        return False
    
    try:
        cursor = conn.cursor()
        table, col = source_table.split('.')
        
        # Insert just the one column - others will be NULL or use defaults
        query = f"INSERT INTO {table} ({col}) VALUES (%s)"
        print(f"DEBUG: Inserting lookup value: {query} with {value}")
        cursor.execute(query, (value,))
        conn.commit()
        cursor.close()
        
        print(f"✓ Added {value} to {table}.{col}")
        return True
    except Exception as e:
        print(f"ERROR adding lookup value: {e}")
        import traceback
        traceback.print_exc()
        return False
    finally:
        return_db_connection(conn)


def generate_lookups_for_table(table_name: str) -> dict:
    """
    Scan table for TEXT columns and return lookup entries.
    Returns: {"column1": "table.column1", "column2": "table.column2", ...}
    """
    try:
        col_types = get_column_types(table_name)
        text_cols = [col for col, col_type in col_types.items() if col_type == 'text']
        
        # Build lookups dict: col -> table.col
        lookups = {col: f"{table_name}.{col}" for col in text_cols}
        return lookups
    except Exception as e:
        print(f"ERROR generating lookups for {table_name}: {e}")
        import traceback
        traceback.print_exc()
        return {}


# ============================================================================
# USER PERMISSION FUNCTIONS
# ============================================================================

def get_user_permissions(username: str) -> dict:
    """
    Get user permissions from fastapi_users table.
    Returns dict: {'view': bool, 'edit': bool, 'add': bool, 'delete': bool, 'admin': bool, 'run_scripts': bool, 'export_data': bool, 'import_data': bool}
    """
    from config import FASTAPI_USERS_TABLE
    
    conn = get_db_connection()
    if not conn:
        print(f"ERROR: Could not connect to database for user permissions")
        return {'view': False, 'edit': False, 'add': False, 'delete': False, 'admin': False, 'run_scripts': False, 'export_data': False, 'import_data': False}
    
    try:
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        query = f'SELECT view, edit, add, delete, admin, run_scripts, export_data, import_data FROM {FASTAPI_USERS_TABLE} WHERE "user" = %s'
        print(f"DEBUG: Fetching permissions for user '{username}': {query}")
        cursor.execute(query, (username,))
        result = cursor.fetchone()
        cursor.close()
        
        if result:
            return {
                'view': result.get('view', False),
                'edit': result.get('edit', False),
                'add': result.get('add', False),
                'delete': result.get('delete', False),
                'admin': result.get('admin', False),
                'run_scripts': result.get('run_scripts', False),
                'export_data': result.get('export_data', False),
                'import_data': result.get('import_data', False),
                'username': username
            }
        
        # User not found
        print(f"WARNING: User '{username}' not found in {FASTAPI_USERS_TABLE}")
        return {'view': False, 'edit': False, 'add': False, 'delete': False, 'admin': False, 'run_scripts': False, 'export_data': False, 'import_data': False, 'username': username}
    
    except Exception as e:
        print(f"ERROR fetching user permissions for '{username}': {e}")
        import traceback
        traceback.print_exc()
        return {'view': False, 'edit': False, 'add': False, 'delete': False, 'admin': False, 'username': username}
    finally:
        return_db_connection(conn)


def ensure_user_exists(username: str) -> bool:
    """
    Check if user exists in fastapi_users table.
    If not, create them with permissions:
    - First user: full permissions (view=true, edit=true, add=true, delete=true, admin=true)
    - Subsequent users: default permissions (view=true, rest=false)
    Returns True if user exists or was created, False on error.
    """
    from config import FASTAPI_USERS_TABLE
    
    conn = get_db_connection()
    if not conn:
        print(f"ERROR: Could not connect to database to ensure user exists")
        return False
    
    try:
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # Check if user already exists
        check_query = f'SELECT COUNT(*) as cnt FROM {FASTAPI_USERS_TABLE} WHERE "user" = %s'
        cursor.execute(check_query, (username,))
        result = cursor.fetchone()
        user_count = result['cnt'] if result else 0
        
        if user_count > 0:
            print(f"✓ User '{username}' already exists in {FASTAPI_USERS_TABLE}")
            cursor.close()
            return_db_connection(conn)
        total_result = cursor.fetchone()
        total_users = total_result['cnt'] if total_result else 0
        
        is_first_user = total_users == 0
        
        if is_first_user:
            # First user gets full permissions
            print(f"DEBUG: First user '{username}' - granting full permissions")
            insert_query = f'INSERT INTO {FASTAPI_USERS_TABLE} ("user", view, edit, add, delete, admin, run_scripts, export_data, import_data) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)'
            cursor.execute(insert_query, (username, True, True, True, True, True, True, True, True))
            conn.commit()
            print(f"✓ Created first user '{username}' with FULL permissions (admin)")
        else:
            # Subsequent users get default permissions
            print(f"DEBUG: Subsequent user '{username}' - granting default permissions")
            insert_query = f'INSERT INTO {FASTAPI_USERS_TABLE} ("user", view, edit, add, delete, admin, run_scripts, export_data, import_data) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)'
            cursor.execute(insert_query, (username, True, False, False, False, False, False, False, False))
            conn.commit()
            print(f"✓ Created new user '{username}' with default permissions (view only)")
        
        cursor.close()
        return_db_connection(conn)
        return True
        
    except Exception as e:
        print(f"ERROR in ensure_user_exists for '{username}': {e}")
        import traceback
        traceback.print_exc()
        try:
            return_db_connection(conn)
        except:
            pass
        return False


# ============================================================================
# IMPORT/EXPORT FUNCTIONS
# ============================================================================

def export_table_to_csv(table_name: str) -> bytes:
    """
    Export all rows from a table to CSV format.
    Returns CSV bytes that can be downloaded.
    """
    import csv
    import io
    
    conn = get_db_connection()
    if not conn:
        raise Exception("Could not connect to database")
    
    try:
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # Get column names
        cursor.execute(
            "SELECT column_name FROM information_schema.columns WHERE table_name = %s ORDER BY ordinal_position",
            (table_name,)
        )
        columns = [row['column_name'] for row in cursor.fetchall()]
        
        # Get all data
        cursor.execute(f"SELECT * FROM {table_name} ORDER BY {columns[0]}")
        rows = cursor.fetchall()
        
        # Write to CSV
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=columns)
        writer.writeheader()
        
        for row in rows:
            # Convert row dict, handling None values
            row_dict = {col: (row[col] if row[col] is not None else '') for col in columns}
            writer.writerow(row_dict)
        
        cursor.close()
        return output.getvalue().encode('utf-8')
    
    finally:
        return_db_connection(conn)


def export_table_to_excel(table_name: str) -> bytes:
    """
    Export all rows from a table to Excel (.xlsx) format.
    Returns Excel bytes that can be downloaded.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font
    import io
    
    conn = get_db_connection()
    if not conn:
        raise Exception("Could not connect to database")
    
    try:
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # Get column names
        cursor.execute(
            "SELECT column_name FROM information_schema.columns WHERE table_name = %s ORDER BY ordinal_position",
            (table_name,)
        )
        columns = [row['column_name'] for row in cursor.fetchall()]
        
        # Get all data
        cursor.execute(f"SELECT * FROM {table_name} ORDER BY {columns[0]}")
        rows = cursor.fetchall()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = table_name[:31]  # Excel sheet name limit is 31 chars
        
        # Write header
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = col_name
            cell.font = Font(bold=True)
        
        # Write data
        for row_idx, row in enumerate(rows, 2):
            for col_idx, col_name in enumerate(columns, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = row[col_name]
        
        # Auto-adjust column widths
        for col_idx, col_name in enumerate(columns, 1):
            ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = min(50, max(12, len(col_name) + 2))
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        cursor.close()
        output.seek(0)
        return output.getvalue()
    
    finally:
        return_db_connection(conn)


def validate_import_row(table_name: str, row_dict: dict) -> tuple:
    """
    Validate a single row for import.
    Returns: (is_valid: bool, errors: list, warnings: list)
    """
    col_types = get_column_types(table_name)
    errors = []
    warnings = []
    
    for col_name, col_type in col_types.items():
        value = row_dict.get(col_name)
        
        # Skip None/empty values
        if value is None or value == '':
            warnings.append(f"{col_name}: empty value")
            continue
        
        # Validate data type
        try:
            if col_type in ['integer', 'bigint', 'smallint']:
                int(value)
            elif col_type in ['numeric', 'decimal', 'double precision', 'real']:
                float(value)
            elif col_type in ['date']:
                # Basic date format check YYYY-MM-DD
                if len(str(value)) == 10 and str(value)[4] == '-' and str(value)[7] == '-':
                    pass  # Format looks ok
                else:
                    warnings.append(f"{col_name}: date may not be in YYYY-MM-DD format")
            elif col_type in ['boolean']:
                # Accept various boolean representations
                if str(value).lower() not in ['true', 'false', '1', '0', 'yes', 'no', 'on', 'off']:
                    warnings.append(f"{col_name}: boolean value may not be recognized")
        except (ValueError, TypeError) as e:
            errors.append(f"{col_name}: invalid {col_type} value ({value})")
    
    is_valid = len(errors) == 0
    return is_valid, errors, warnings


def insert_imported_rows(table_name: str, rows: list, skip_duplicates: bool = True, dry_run: bool = False, replace_all: bool = False) -> dict:
    """
    Insert rows into table with validation and duplicate handling.
    rows: list of dicts with column_name: value pairs
    skip_duplicates: if True, skip rows with duplicate primary keys
    replace_all: if True, delete all existing data first and insert all rows
    dry_run: if True, validate but don't insert
    
    Returns: {
        'inserted': int,
        'skipped': int,
        'errors': list of error messages,
        'warnings': list of warning messages
    }
    """
    import copy
    
    conn = get_db_connection()
    if not conn:
        return {'inserted': 0, 'skipped': 0, 'errors': ['Could not connect to database'], 'warnings': []}
    
    try:
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # Get columns
        cursor.execute(
            "SELECT column_name FROM information_schema.columns WHERE table_name = %s ORDER BY ordinal_position",
            (table_name,)
        )
        columns = [row['column_name'] for row in cursor.fetchall()]
        
        # Get primary key column using information_schema
        cursor.execute(
            """SELECT a.attname 
               FROM information_schema.table_constraints tc
               JOIN information_schema.key_column_usage kcu ON tc.constraint_name = kcu.constraint_name
               JOIN pg_attribute a ON a.attname = kcu.column_name
               WHERE tc.table_name = %s AND tc.constraint_type = 'PRIMARY KEY'
               LIMIT 1""",
            (table_name,)
        )
        pk_result = cursor.fetchone()
        pk_column = pk_result['attname'] if pk_result else columns[0]  # Default to first column
        
        inserted = 0
        skipped = 0
        all_errors = []
        all_warnings = []
        
        # Truncate table if replace_all is True
        if replace_all and not dry_run:
            try:
                cursor.execute(f"TRUNCATE TABLE {table_name} RESTART IDENTITY CASCADE")
            except Exception as e:
                all_errors.append(f"Failed to truncate table: {str(e)}")
                conn.rollback()
                return {'inserted': 0, 'skipped': 0, 'errors': all_errors, 'warnings': []}
        
        for row_idx, row in enumerate(rows, 1):
            # Validate row
            is_valid, errors, warnings = validate_import_row(table_name, row)
            
            if errors:
                all_errors.append(f"Row {row_idx}: {', '.join(errors)}")
                continue
            
            all_warnings.extend([f"Row {row_idx}: {w}" for w in warnings])
            
            # Check for duplicates (skip if replace_all is True since we truncated everything)
            if not replace_all and skip_duplicates and pk_column in row:
                pk_value = row[pk_column]
                cursor.execute(f"SELECT COUNT(*) as cnt FROM {table_name} WHERE {pk_column} = %s", (pk_value,))
                exists = cursor.fetchone()['cnt'] > 0
                if exists:
                    skipped += 1
                    continue
            
            # Insert row (skip if dry_run)
            if not dry_run:
                # Create savepoint for this row so one error doesn't abort all remaining rows
                savepoint_name = f"sp_{row_idx}"
                try:
                    cursor.execute(f"SAVEPOINT {savepoint_name}")
                    
                    # Filter out columns not in table
                    insert_dict = {col: row.get(col) for col in columns if col in row}
                    col_names = list(insert_dict.keys())
                    col_values = [insert_dict[col] for col in col_names]
                    
                    placeholders = ', '.join(['%s'] * len(col_names))
                    col_names_str = ', '.join([f'"{col}"' for col in col_names])
                    
                    # Build INSERT or UPSERT query
                    if replace_all:
                        # Simple INSERT - table already truncated
                        insert_query = f"INSERT INTO {table_name} ({col_names_str}) VALUES ({placeholders})"
                    elif skip_duplicates:
                        # Simple INSERT - skip on duplicate
                        insert_query = f"INSERT INTO {table_name} ({col_names_str}) VALUES ({placeholders})"
                    else:
                        # UPSERT - update existing rows, insert new ones
                        # Update all columns except the primary key
                        update_cols = [col for col in col_names if col != pk_column]
                        if update_cols:
                            update_set = ', '.join([f'"{col}" = EXCLUDED."{col}"' for col in update_cols])
                            insert_query = f"INSERT INTO {table_name} ({col_names_str}) VALUES ({placeholders}) ON CONFLICT (\"{pk_column}\") DO UPDATE SET {update_set}"
                        else:
                            insert_query = f"INSERT INTO {table_name} ({col_names_str}) VALUES ({placeholders}) ON CONFLICT (\"{pk_column}\") DO NOTHING"
                    
                    cursor.execute(insert_query, col_values)
                    inserted += 1
                except Exception as e:
                    # Rollback to savepoint to continue with next row
                    cursor.execute(f"ROLLBACK TO SAVEPOINT {savepoint_name}")
                    
                    # Handle duplicate key errors if skip_duplicates is true
                    error_msg = str(e)
                    if skip_duplicates and ('duplicate key' in error_msg.lower() or 'unique constraint' in error_msg.lower()):
                        skipped += 1
                    else:
                        all_errors.append(f"Row {row_idx}: Insert failed - {error_msg}")
            else:
                inserted += 1  # Count as inserted in dry-run
        
        if not dry_run:
            conn.commit()
        
        cursor.close()
        return {
            'inserted': inserted,
            'skipped': skipped,
            'errors': all_errors,
            'warnings': all_warnings
        }
    
    finally:
        return_db_connection(conn)
